from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from time import sleep

from openpyxl import Workbook, load_workbook


def load_categories(file_name):
    categories_wb = load_workbook("categories.xlsx", read_only=True)
    categories_ws = categories_wb.active
    categories_info = [[cell.value for cell in sub_category ] for sub_category in categories_ws.rows]
    return categories_info

def save_info(info, file_name):
    wb = Workbook()
    ws = wb.active
    for row in info:
        ws.append(row)
    wb.save(file_name)   

def get_products_from_url(browser, category_name, sub_category_name, sub_category_url):
    products_info = []
    browser.get(sub_category_url + "?page=200")
    sleep(2)
    cards = browser.find_elements_by_class_name("plp-fragment-wrapper")
    for card in cards:
        product_name = card.find_element_by_class_name("range-revamp-header-section__title--small").text
        product_description = card.find_element_by_class_name("range-revamp-header-section__description").text
        product_price = card.find_element_by_class_name("range-revamp-price__integer").text
        product_url = card.find_element_by_tag_name("a").get_attribute('href')
        product_code = product_url.split("-")[-1].strip("s/")
        products_images = [img_tag.get_attribute('src') for img_tag in card.find_elements_by_tag_name("img")]
        product_info = [category_name, sub_category_name, product_name, 
                        product_description, product_price,
                        product_code, len(products_images)]
        product_info += products_images
        products_info.append(product_info)
    return products_info

if __name__ == "__main__":
    path_to_chromedriver = "./chromedriver.exe"
    browser = webdriver.Chrome(executable_path = path_to_chromedriver)
    categories_info = load_categories("categoreis.xlsx")
    products_info = []
    i = 1
    for sub_category in categories_info:
        print(f"descargando de {sub_category[1]} ({i}/{len(categories_info)})")
        products_info += get_products_from_url(browser, sub_category[0], sub_category[1], sub_category[2])
        i += 1
    save_info(products_info, "products_catalog.xlsx")
    browser.close()